import math
import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from CONFIG import *

# открытие браузера Chrome
options = Options()
# options.add_argument('--headless')
# options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument(f"user-agent={get_random_user_agent()}")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
wait = WebDriverWait(driver, 12)

# ищем оквэд по имени и индексу
def search_okved_by_name_and_index(company_name: str, company_index: str) -> str:
    OKVED = ""

    # переход на сайт
    # wait_random()
    driver.get(url)

    if ("403 Forbidden" or "Соединение прервано") in driver.page_source:
        debug("Ошибка загрузки страницы")
        input("\033[38;2;255;140;0m[ВНИМАНИЕ] ОБНАРУЖЕНА ОШИБКА ЗАГРУЗКИ! ПОЧИНИТЕ И НАЖМИТЕ ENTER: \033[0m")

    if ("«Я не робот»." or "вы человек") in driver.page_source:
        clicker_vpn()
        driver.get(url)

        # time.sleep(20)

        try:
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            if ("«Я не робот»." or "вы человек" or "Forbidden") in driver.page_source:
                input("\033[38;2;255;140;0m[ВНИМАНИЕ] ОБНАРУЖЕНА КАПЧА! ВВЕДИТЕ КАПЧУ И НАЖМИТЕ ENTER: \033[0m")
        except:
            debug("Ошибка загрузки страницы")
            input("\033[38;2;255;140;0m[ВНИМАНИЕ] ОБНАРУЖЕНА ОШИБКА ЗАГРУЗКИ! ПОЧИНИТЕ И НАЖМИТЕ ENTER: \033[0m")

    try:
        # находим поле input и очищаем его
        search_input = driver.find_element(By.CSS_SELECTOR, "input.index-search-input")
        search_input.clear()

        # вводим название компании и нажимаем кнопку поиска
        search_input.send_keys(company_name)
        search_input.send_keys(Keys.RETURN)
        # print(driver.page_source)

        # проверяем, если страниц несколько, то идем по всем page-nav
        count_pages = 1
        try:
            if driver.find_elements(By.CSS_SELECTOR, "div.page-nav"):
                try:
                    if not driver.find_element(By.CSS_SELECTOR, "span.page-navigation__num") is None:
                        count_pages = math.ceil(int(driver.find_element(By.CSS_SELECTOR, "span.page-navigation__num").text[3:]) / 100)
                        if count_pages > 10:
                            count_pages = 10
                except ValueError:
                    debug("int(driver.find_element(By.CSS_SELECTOR, \"span.page-navigation__num\").text)")
        except ValueError:
            debug("driver.find_element(By.CSS_SELECTOR, \"div.page-nav\")")

        debug(f"count_pages = {count_pages}")

        for count_page in range(count_pages):
            try:

                # если страница такая одна
                try:
                    company_okv_data = driver.find_elements(By.CSS_SELECTOR, "span.bolder")
                    if len(company_okv_data) > 0:
                        for okv in OKVEDS_TRUE:
                            if okv in company_okv_data[0].text:
                                OKVED = okv
                except ValueError:
                    debug("driver.find_elements(By.CSS_SELECTOR, \"span.bolder\")")

                # находим все div с class="company-item"
                company_items = driver.find_elements(By.CSS_SELECTOR, "div.company-item")
                for item in company_items:
                    # Найти адреса всех компаний
                    try:
                        company_index_text = ""
                        try:
                            count_company_index_text = item.find_elements(By.CSS_SELECTOR, "address.company-item__text")
                            if len(count_company_index_text) > 0:
                                company_index_text = count_company_index_text[0].text
                        except ValueError:
                            debug("item.find_element(By.CSS_SELECTOR, \"address.company-item__text\")")

                        if company_index in company_index_text:
                            try:
                                item_info = item.find_elements(By.CSS_SELECTOR, "div.company-item-info")
                                # Найти элемент "Основной вид деятельности"
                                if len(item_info) >= 3:
                                    for okv in OKVEDS_TRUE:
                                        if (okv+" ") in item_info[2].text:
                                            OKVED = okv
                            except ValueError:
                                debug("item.find_elements(By.CSS_SELECTOR, \"div.company-item-info\")")
                    except ValueError:
                        debug("except item.find_element(By.CSS_SELECTOR, \"address.company-item__text\").text")
                if 1 < count_pages != count_page+1:
                    # находим кнопку далее и жмем по ней
                    try:
                        if not driver.find_element(By.CSS_SELECTOR, "a.nav-next") is None:
                            driver.find_element(By.CSS_SELECTOR, "a.nav-next").click()
                    except ValueError:
                        debug("except driver.find_element(By.CSS_SELECTOR, \"a.nav-next\").click()")
            except ValueError:
                debug("except driver.find_elements(By.CSS_SELECTOR, \"div.company-item\")")

            if OKVED != "":
                break

    except ValueError:
        debug("except driver.find_element(By.CSS_SELECTOR, \"input.index-search-input\")")

    return OKVED

# РАБОТА С ТАБЛИЦЕЙ

# открываем файл и выбираем нужную таблицу
workbook = load_workbook(filename=path_to_file)
sheet = workbook.active

# пропускаем первую строку с заголовками
for row in sheet.iter_rows(min_row=progress_bar+1):
    # чтобы понимать, где я нахожусь
    debug("\033[31m[PROGRESS BAR = " + str(progress_bar) + "]\033[0m")
    write_progress(progress_bar)

    # получаем нужные значения ячеек
    company_name = str(row[1].value)
    company_index = str(row[6].value)

    if company_name is None or company_name == "":
        debug("SKIPPED")
        progress_bar += 1
        break

    # вызываем функцию поиска
    buffer = search_okved_by_name_and_index(company_name, company_index)

    # отладка поиска оквэда
    debug(f"{company_name} $ {company_index} = " + buffer)

    # записываем результат в нужную ячейку
    row[19].value = buffer

    # сохраняем изменения в файле
    workbook.save(path_to_file)
    debug(f"workbook.save(\"{path_to_file}\")")

    # каждые 100 записей делаем копию
    if progress_bar % copy_const == 0:
        backup_file(progress_bar, path_to_file)
        options.add_argument(f"user-agent={get_random_user_agent()}")
        debug("\033[38;2;0;255;0m[BACKUP & CHANGE USER-AGENT]\033[0m")

    # чтобы понимать, где я нахожусь
    progress_bar += 1

# закрытие браузера
debug("driver.close()")
# time.sleep(5)
driver.close()